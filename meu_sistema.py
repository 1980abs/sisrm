import streamlit as st
import pandas as pd
import io
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# Configuração da página
st.set_page_config(page_title="Sistema de Afiliados", page_icon="📊", layout="centered")

# --- NOVIDADE: Criação de uma "memória" para o botão de limpar funcionar ---
if "chave_uploader" not in st.session_state:
    st.session_state.chave_uploader = 0

def limpar_dados():
    st.session_state.chave_uploader += 1

# Adiciona a logo que já está no seu GitHub
if os.path.exists("logo.png"):
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.image("logo.png", use_column_width=True)

st.title("📊 Comparador de Planilhas - Afiliados")
st.write("Faça o upload de uma ou mais planilhas (Excel ou CSV) para encontrar Logins duplicados. Os duplicados serão agrupados no topo e pintados de Azul Escuro.")

# Botão para limpar os dados (Fica no topo para facilitar)
st.button("🔄 Limpar Dados / Iniciar Nova Análise", on_click=limpar_dados)

# Área para enviar os arquivos (agora conectada à nossa "memória" para poder resetar)
arquivos_enviados = st.file_uploader(
    "Arraste e solte as planilhas aqui", 
    type=['xlsx', 'csv'], 
    accept_multiple_files=True,
    key=str(st.session_state.chave_uploader)
)

if arquivos_enviados:
    st.success(f"{len(arquivos_enviados)} arquivo(s) carregado(s) com sucesso!")
    
    if st.button("🚀 Analisar e Juntar Planilhas"):
        with st.spinner('Analisando os dados...'):
            lista_planilhas = []
            
            # Lê todos os arquivos enviados
            for arquivo in arquivos_enviados:
                if arquivo.name.endswith('.csv'):
                    df = pd.read_csv(arquivo)
                else:
                    df = pd.read_excel(arquivo)
                
                # Adiciona uma coluna para sabermos de onde veio o dado
                df['Origem_Arquivo'] = arquivo.name 
                lista_planilhas.append(df)
            
            # Junta todas as planilhas em uma só
            planilha_completa = pd.concat(lista_planilhas, ignore_index=True)
            
            # Remove colunas totalmente vazias (sujeira do excel)
            planilha_completa = planilha_completa.dropna(axis=1, how='all')
            
            # Procura a coluna 'Login'
            coluna_login = None
            for col in planilha_completa.columns:
                if str(col).strip().lower() == 'login':
                    coluna_login = col
                    break
            
            if not coluna_login:
                st.error("Erro: Não encontrei nenhuma coluna chamada 'Login' nas planilhas.")
            else:
                # Limpa espaços em branco para a comparação ser exata
                planilha_completa[coluna_login] = planilha_completa[coluna_login].astype(str).str.strip()
                
                # Cria uma coluna temporária para identificar os duplicados
                planilha_completa['Eh_Duplicado'] = planilha_completa.duplicated(subset=[coluna_login], keep=False)
                
                # Ordena jogando os duplicados para o topo e em ordem alfabética
                planilha_completa = planilha_completa.sort_values(
                    by=['Eh_Duplicado', coluna_login], 
                    ascending=[False, True]
                ).reset_index(drop=True)
                
                qtd_duplicados = planilha_completa['Eh_Duplicado'].sum()
                
                # Remove a coluna temporária
                planilha_completa = planilha_completa.drop(columns=['Eh_Duplicado'])
                
                # Recalcula a máscara de duplicados na nova ordem
                duplicados_mascara = planilha_completa.duplicated(subset=[coluna_login], keep=False)
                
                st.warning(f"Análise concluída! Encontramos **{qtd_duplicados}** linhas com Logins repetidos.")
                
                # Mostra uma prévia na tela
                st.write("Prévia dos dados (Duplicados agrupados no topo):")
                st.dataframe(planilha_completa.head(10))
                
                # --- PARTE DE PINTAR A PLANILHA DE AZUL ---
                buffer = io.BytesIO()
                planilha_completa.to_excel(buffer, index=False, engine='openpyxl')
                buffer.seek(0)
                
                wb = load_workbook(buffer)
                ws = wb.active
                
                # Configura a cor Azul Escuro e a Letra Branca
                fundo_azul_escuro = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")
                fonte_branca = Font(color="FFFFFF", bold=True)
                
                # Acha o número da coluna Login no Excel gerado
                idx_coluna_login = None
                for idx, celula in enumerate(ws[1], start=1):
                    if celula.value == coluna_login:
                        idx_coluna_login = idx
                        break
                        
                if idx_coluna_login:
                    # Pinta as células duplicadas
                    for num_linha, eh_duplicado in enumerate(duplicados_mascara, start=2):
                        if eh_duplicado:
                            celula = ws.cell(row=num_linha, column=idx_coluna_login)
                            celula.fill = fundo_azul_escuro
                            celula.font = fonte_branca
                            
                # Salva o arquivo pintado na memória
                buffer_final = io.BytesIO()
                wb.save(buffer_final)
                buffer_final.seek(0)
                
                # Cria o botão de Download
                st.success("Tudo pronto! Sua planilha colorida e organizada já pode ser baixada.")
                st.download_button(
                    label="📥 Baixar Planilha Final (Organizada e com duplicados em Azul)",
                    data=buffer_final,
                    file_name="Planilha_Comparada_Organizada.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )